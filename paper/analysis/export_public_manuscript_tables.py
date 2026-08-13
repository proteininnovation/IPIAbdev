#!/usr/bin/env python3
"""Build public DELPHI manuscript tables without literal IPI sequences."""
from __future__ import annotations

import hashlib
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

PAPER = Path(__file__).resolve().parents[1]
PACKAGE = PAPER.parent
FINAL = PACKAGE.parent / "manuscript_update_20260810" / "final_submission" / "final"
REPO = Path(__file__).resolve().parents[5]
OUTPUT = PAPER / "data" / "shareable" / "manuscript_tables"

SUPPLEMENTARY = {
    2: FINAL / "DELPHI_Supplementary_Table_2.xlsx",
    4: FINAL / "DELPHI_Supplementary_Table_4.xlsx",
    5: FINAL / "DELPHI_Supplementary_Table_5.xlsx",
    7: FINAL / "DELPHI_Supplementary_Table_7.xlsx",
}
EXTENDED = {
    1: REPO / "Final" / "data" / "ED_Table1_PSR.xlsx",
    2: REPO / "Final" / "data" / "ED_Table2_PSR.xlsx",
    3: REPO / "Final" / "data" / "ED_Table3_SEC.xlsx",
}

FORBIDDEN_HEADERS = {
    "HSEQ", "LSEQ", "CDR3", "HCDR3", "VH_SEQUENCE", "VL_SEQUENCE",
}
SEQUENCE_RE = re.compile(r"(?:Biotin_)?C[ACDEFGHIKLMNPQRSTVWY]{8,30}(?=[:_\s,<]|$)")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def audit(path: Path, public_sequence_sheets: set[str] | None = None) -> dict:
    public_sequence_sheets = public_sequence_sheets or set()
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    errors = []
    for sheet in workbook.worksheets:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        bad = sorted(FORBIDDEN_HEADERS.intersection(value.upper() for value in headers))
        if bad and sheet.title not in public_sequence_sheets:
            errors.append(f"{sheet.title}: forbidden headers {bad}")
        sequence_hits = 0
        formula_count = 0
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    formula_count += int(value.startswith("="))
                    sequence_hits += int(SEQUENCE_RE.search(value) is not None)
        if sequence_hits and sheet.title not in public_sequence_sheets:
            errors.append(f"{sheet.title}: {sequence_hits} sequence-like cell values")
        sheets.append({
            "name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "formula_cells": formula_count,
            "sequence_like_cells": sequence_hits,
        })
    return {"sheets": sheets, "errors": errors}


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    manifest = {
        "policy": "No private IPI antibody sequences. Public DS1/Jain2017/GDPa cohort sequences are retained in the official supplementary tables.",
        "files": [],
    }

    for number, source in SUPPLEMENTARY.items():
        destination = OUTPUT / f"DELPHI_Supplementary_Table_{number}.xlsx"
        shutil.copy2(source, destination)
        allowed = {"DS1"} if number == 4 else ({"Jain2017", "GDPa1", "GDPa3"} if number == 5 else set())
        result = audit(destination, public_sequence_sheets=allowed)
        manifest["files"].append({
            "file": destination.name,
            "source": source.name,
            "exact_copy_of_final_submission": sha256(source) == sha256(destination),
            "public_sequence_sheets": sorted(allowed),
            "sha256": sha256(destination),
            **result,
        })

    for number, source in EXTENDED.items():
        destination = OUTPUT / f"DELPHI_Extended_Data_Table_{number}.xlsx"
        shutil.copy2(source, destination)
        result = audit(destination)
        manifest["files"].append({
            "file": destination.name,
            "source": source.name,
            "exact_copy_of_source": sha256(source) == sha256(destination),
            "public_sequence_sheets": [],
            "sha256": sha256(destination),
            **result,
        })

    errors = [f"{item['file']}: {error}" for item in manifest["files"] for error in item["errors"]]
    manifest["audit"] = "passed" if not errors else "failed"
    manifest["audit_errors"] = errors
    (OUTPUT / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    if errors:
        raise SystemExit("Public-table sequence audit failed:\n- " + "\n- ".join(errors))
    print(f"Exported {len(manifest['files'])} public manuscript tables to {OUTPUT}")
    print("Sequence audit: PASSED")


if __name__ == "__main__":
    main()
